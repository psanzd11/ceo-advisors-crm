#!/usr/bin/env python3
"""
CEO Advisors CRM — Data Injection Script
Schema version: 2026.05.1

Usage:
  python inject_data.py                       # Modo seed completo (sobrescribe seedData)
  python inject_data.py --merge --from CRM_export.json  # Modo merge — solo IDs nuevos

Lee CEO_Advisors_CRM_DataTemplate.xlsx y:
  - Ejecuta validación pre-flight (referential integrity + splits sum + duplicados).
  - En modo seed: inyecta los datos como semilla en CEO_Advisors_CRM.html → CEO_Advisors_CRM_PRODUCTION.html
  - En modo merge: genera merge_patch.json con solo los IDs no presentes en el export del CRM.
"""

import json, re, sys, argparse, hashlib, secrets, base64
from pathlib import Path

# ── PASSWORD HASHING (PBKDF2-HMAC-SHA256) ────────────────
# Compatible con SubtleCrypto.deriveBits('PBKDF2') del navegador.
PBKDF2_ITERS = 120000  # ~150-300ms en máquina moderna; suficiente para CRM interno
PBKDF2_KEYLEN = 32     # 256 bits
WEAK_PASSWORDS = {"changeme", "demo1234", "pablo1234", "admin1234", "securepass1!", "password", "1234", ""}

def hash_password(plain):
    """Devuelve {hash, salt, iters} en base64. Si la password es 'demo' (vacía o débil),
    marca passwordMustChange=True para forzar cambio en primer login."""
    salt = secrets.token_bytes(16)
    pw = (plain or "changeme").encode("utf-8")
    digest = hashlib.pbkdf2_hmac("sha256", pw, salt, PBKDF2_ITERS, PBKDF2_KEYLEN)
    return {
        "passwordHash": base64.b64encode(digest).decode("ascii"),
        "passwordSalt": base64.b64encode(salt).decode("ascii"),
        "passwordIters": PBKDF2_ITERS,
        "passwordMustChange": (plain or "").strip().lower() in WEAK_PASSWORDS,
    }

# ── SCHEMA VERSION ────────────────────────────────────
SCHEMA_VERSION = "2026.05.1"
SCHEMA_FIELDS = {
    "consultants": ["id","name","role","isCEO","isAdmin","email","password","region","bio"],
    "clients":     ["id","name","email","phone","country","city","netWorth","source","tier","notes"],
    "companies":   ["id","name","industry","country","employees","netWorth","website","clientIds","notes"],
    "deals":       ["id","title","clientId","companyId","type","stage","amount","closeDate","createdAt","splits","notes"],
    "activities":  ["id","type","date","clientId","companyId","dealId","title","notes","done","createdBy"],
    "pupilos":     ["id","name","email","university","program","startDate","endDate","mentor","region","consultantId","leftCompany","leftRole","notes"],
}

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl..."); import subprocess
    subprocess.run([sys.executable,"-m","pip","install","openpyxl","--break-system-packages","-q"])
    import openpyxl

HERE = Path(__file__).parent
# Si _v2.NEW.xlsx existe (porque sync.py no pudo guardar el original por estar abierto en Excel),
# lo usamos preferentemente para no perder los cambios sincronizados.
_v2_new = HERE / "CEO_Advisors_CRM_DataTemplate_v2.NEW.xlsx"
_v2 = HERE / "CEO_Advisors_CRM_DataTemplate_v2.xlsx"
TEMPLATE_XL = _v2_new if _v2_new.exists() else _v2
CRM_HTML    = HERE / "CEO_Advisors_CRM.html"
OUTPUT_HTML = HERE / "CEO_Advisors_CRM_PRODUCTION.html"
MERGE_OUT   = HERE / "merge_patch.json"
PUPILO_DOCS_DIR = "pupilo_docs"  # carpeta relativa al HTML donde viven los CVs

# ── HELPERS ───────────────────────────────────────────
def v(cell):
    val = cell.value
    if val is None: return ""
    return str(val).strip()

def num(cell):
    val = cell.value
    if val is None: return 0
    try: return int(float(str(val).replace(",","")))
    except: return 0

def pct(cell):
    val = cell.value
    if val is None: return 0
    try: return int(float(str(val).replace("%","").strip()))
    except: return 0

def bool_cell(cell):
    return v(cell).lower() in ("yes","sí","si","true","1","done","x")

def check_schema_version(ws, sheet_name):
    """Lee la versión declarada en A1 (si existe) y compara."""
    raw = v(ws.cell(1,1))
    m = re.search(r"v(\d{4}\.\d{2}\.\d+)", raw)
    if not m:
        print(f"  ⚠ {sheet_name}: A1 no declara versión de schema.")
        return
    if m.group(1) != SCHEMA_VERSION:
        print(f"  ⚠ {sheet_name}: schema {m.group(1)} ≠ script {SCHEMA_VERSION}. Revisa columnas.")

# ── PARSE EXCEL ───────────────────────────────────────
def parse_workbook():
    wb = openpyxl.load_workbook(TEMPLATE_XL, data_only=True)

    consultants, clients, companies, deals, activities, pupilos = [], [], [], [], [], []

    def auto_id(row_cell, prefix, idx):
        """Devuelve el ID. Si la columna A está vacía (fórmula sin calcular),
        usa el prefijo + índice de fila relativo a la primera fila de datos."""
        explicit = v(row_cell)
        return explicit if explicit else f"{prefix}{idx}"

    # Consultants
    if "👥 Consultants" in wb.sheetnames:
        ws = wb["👥 Consultants"]; check_schema_version(ws, "Consultants")
        idx = 0
        for row in ws.iter_rows(min_row=4):
            name = v(row[1])
            if not name: continue
            idx += 1
            uid = auto_id(row[0], "u", idx)
            consultants.append({
                "id":uid,"name":name,"role":v(row[2]) or "Advisor",
                "isCEO":bool_cell(row[3]),
                "isAdmin":bool_cell(row[4]) or bool_cell(row[3]),
                "email":v(row[5]),"password":v(row[6]) or "changeme",
                "region":v(row[7]),"bio":v(row[8]),
            })

    # Clients (10 columnas tras eliminar Assigned Advisor)
    if "🤝 Clients" in wb.sheetnames:
        ws = wb["🤝 Clients"]; check_schema_version(ws, "Clients")
        idx = 0
        for row in ws.iter_rows(min_row=4):
            name = v(row[1])
            if not name: continue
            idx += 1
            cid = auto_id(row[0], "c", idx)
            clients.append({
                "id":cid,"name":name,"email":v(row[2]),"phone":v(row[3]),
                "country":v(row[4]),"city":v(row[5]),"netWorth":num(row[6]),
                "source":v(row[7]) or "Other","tier":v(row[8]),
                "notes":v(row[9]),
            })

    # Companies
    if "🏢 Companies" in wb.sheetnames:
        ws = wb["🏢 Companies"]; check_schema_version(ws, "Companies")
        idx = 0
        for row in ws.iter_rows(min_row=4):
            name = v(row[1])
            if not name: continue
            idx += 1
            coid = auto_id(row[0], "co", idx)
            client_ids = [x.strip() for x in re.split(r"[,;]", v(row[7])) if x.strip()]
            companies.append({
                "id":coid,"name":name,"industry":v(row[2]),"country":v(row[3]),
                "employees":num(row[4]),"netWorth":num(row[5]),"website":v(row[6]),
                "clientIds":client_ids,"notes":v(row[8]),
            })

    # Deals
    if "💼 Deals" in wb.sheetnames:
        ws = wb["💼 Deals"]; check_schema_version(ws, "Deals")
        idx = 0
        for row in ws.iter_rows(min_row=4):
            title = v(row[1])
            if not title: continue
            idx += 1
            did = auto_id(row[0], "d", idx)
            splits = []
            for i in range(3):
                aid = v(row[9 + i*2]); ap = pct(row[10 + i*2])
                if aid: splits.append({"u":aid,"pct":ap})
            deals.append({
                "id":did,"title":title,"clientId":v(row[2]),
                "companyId":v(row[3]) or None,"type":v(row[4]) or "advise",
                "stage":v(row[5]) or "prospect","amount":num(row[6]),
                "closeDate":v(row[7]),"createdAt":v(row[8]),
                "splits":splits,"notes":v(row[15]),
            })

    # Activities (nueva pestaña)
    if "📋 Activities" in wb.sheetnames:
        ws = wb["📋 Activities"]; check_schema_version(ws, "Activities")
        idx = 0
        for row in ws.iter_rows(min_row=4):
            title = v(row[6])
            if not title: continue
            idx += 1
            aid = auto_id(row[0], "a", idx)
            activities.append({
                "id":aid,"type":v(row[1]) or "task","date":v(row[2]),
                "clientId":v(row[3]) or None,"companyId":v(row[4]) or None,
                "dealId":v(row[5]) or None,"title":title,
                "notes":v(row[7]),"done":bool_cell(row[8]),
                "createdBy":v(row[9]),
            })

    # Pupilos (nueva pestaña)
    if "🎓 Pupilos" in wb.sheetnames:
        ws = wb["🎓 Pupilos"]; check_schema_version(ws, "Pupilos")
        idx = 0
        for row in ws.iter_rows(min_row=4):
            name = v(row[1])
            if not name: continue
            idx += 1
            pid = auto_id(row[0], "p", idx)
            # CV (col N, 14ª columna) — opcional. Si está, agregamos doc con path relativo.
            cv_file = v(row[13]) if len(row) > 13 else ""
            docs = []
            if cv_file:
                ext = cv_file.rsplit(".", 1)[-1].lower() if "." in cv_file else ""
                mime = {"pdf":"application/pdf","docx":"application/vnd.openxmlformats-officedocument.wordprocessingml.document","doc":"application/msword"}.get(ext,"application/octet-stream")
                docs.append({"name":cv_file,"path":f"{PUPILO_DOCS_DIR}/{cv_file}","kind":"cv","mime":mime,"uploadedAt":"2026-05-06","size":"CV"})
            pupilos.append({
                "id":pid,"name":name,"email":v(row[2]),
                "university":v(row[3]),"program":v(row[4]),
                "startDate":v(row[5]),"endDate":v(row[6]),
                "mentor":v(row[7]) or None,"region":v(row[8]),
                "consultantId":v(row[9]) or None,
                "leftCompany":v(row[10]),"leftRole":v(row[11]),
                "notes":v(row[12]),"docs":docs,
            })

    return consultants, clients, companies, deals, activities, pupilos

# ── VALIDATION ────────────────────────────────────────
def validate(consultants, clients, companies, deals, activities, pupilos):
    errors, warnings = [], []

    def dup_check(items, label):
        seen = {}
        for it in items:
            seen.setdefault(it["id"], []).append(it.get("name") or it.get("title") or "?")
        for k, v_ in seen.items():
            if len(v_) > 1:
                errors.append(f"[{label}] ID duplicado '{k}' → usado por: {', '.join(v_)}")

    dup_check(consultants, "Consultants")
    dup_check(clients,     "Clients")
    dup_check(companies,   "Companies")
    dup_check(deals,       "Deals")
    dup_check(activities,  "Activities")
    dup_check(pupilos,     "Pupilos")

    cli_ids = {c["id"] for c in clients}
    co_ids  = {c["id"] for c in companies}
    u_ids   = {c["id"] for c in consultants}

    # Companies → clientIds
    for co in companies:
        for cid in co["clientIds"]:
            if cid not in cli_ids:
                errors.append(f"[Companies/{co['id']}] referencia cliente inexistente: '{cid}'")

    # Deals → client/company/splits
    for d in deals:
        if d["clientId"] and d["clientId"] not in cli_ids:
            errors.append(f"[Deals/{d['id']}] cliente inexistente: '{d['clientId']}'")
        if d["companyId"] and d["companyId"] not in co_ids:
            errors.append(f"[Deals/{d['id']}] empresa inexistente: '{d['companyId']}'")
        for s in d["splits"]:
            if s["u"] not in u_ids:
                errors.append(f"[Deals/{d['id']}] advisor inexistente: '{s['u']}'")
        # Splits sum = 100 (tolerancia 0)
        if d["splits"]:
            total = sum(s["pct"] for s in d["splits"])
            if total != 100:
                errors.append(f"[Deals/{d['id']}] splits suman {total}%, deben sumar 100. Detalle: "
                              + ", ".join(f"{s['u']}={s['pct']}%" for s in d["splits"]))
        else:
            warnings.append(f"[Deals/{d['id']}] sin advisors — se asignará al primer consultor.")

    # Activities → relaciones
    for a in activities:
        if a["clientId"] and a["clientId"] not in cli_ids:
            errors.append(f"[Activities/{a['id']}] cliente inexistente: '{a['clientId']}'")
        if a["companyId"] and a["companyId"] not in co_ids:
            errors.append(f"[Activities/{a['id']}] empresa inexistente: '{a['companyId']}'")
        if a["dealId"] and a["dealId"] not in {d["id"] for d in deals}:
            errors.append(f"[Activities/{a['id']}] deal inexistente: '{a['dealId']}'")
        if a["createdBy"] and a["createdBy"] not in u_ids:
            errors.append(f"[Activities/{a['id']}] createdBy inexistente: '{a['createdBy']}'")

    # Pupilos → relaciones
    for p in pupilos:
        if p["mentor"] and p["mentor"] not in u_ids:
            errors.append(f"[Pupilos/{p['id']}] mentor inexistente: '{p['mentor']}'")
        if p["consultantId"] and p["consultantId"] not in u_ids:
            errors.append(f"[Pupilos/{p['id']}] consultantId inexistente: '{p['consultantId']}' (debe existir como consultor)")

    return errors, warnings

# ── BUILD SEED ────────────────────────────────────────
def build_seed(consultants, clients, companies, deals, activities, pupilos):
    if not consultants:
        print("⚠ No consultants found. Aborting."); sys.exit(1)
    if not [c for c in consultants if c.get("isCEO")]:
        consultants[0]["isCEO"] = True; consultants[0]["isAdmin"] = True
        print(f"  → No CEO found; assigned CEO to {consultants[0]['name']}")
    # Hash de passwords (PBKDF2). El password en texto plano del Excel se elimina del seed.
    weak_count = 0
    for c in consultants:
        plain = c.pop("password", "")
        h = hash_password(plain)
        c.update(h)
        if h["passwordMustChange"]:
            weak_count += 1
    if weak_count:
        print(f"  → {weak_count}/{len(consultants)} consultores tienen password débil/demo (se forzará cambio en primer login).")
    first_id = consultants[0]["id"]
    for d in deals:
        if not d["splits"]:
            d["splits"] = [{"u":first_id,"pct":100}]
    from datetime import datetime as _dt
    return {
        "schemaVersion": SCHEMA_VERSION,
        "_seedTs": _dt.now().isoformat(timespec="seconds"),
        "consultants":consultants,"clients":clients,
        "companies":companies,"deals":deals,
        "activities":activities,
        "pupilos":pupilos,"currentUserId":first_id,
        "audit":[],"notifications":[],
        "winLossReasons":["Pricing","Timing","Competitor","Fit del producto","Cambio decisor","Presupuesto","Otro"],
    }

# ── MERGE MODE ────────────────────────────────────────
def merge_mode(seed, from_path):
    p = Path(from_path)
    if not p.exists():
        print(f"⚠ --from {from_path} no existe. Aborto."); sys.exit(1)
    current = json.loads(p.read_text(encoding="utf-8"))
    patch = {"schemaVersion": SCHEMA_VERSION, "_mode":"merge"}
    for key in ("consultants","clients","companies","deals","activities","pupilos"):
        existing_ids = {x["id"] for x in current.get(key,[])}
        new_only = [x for x in seed[key] if x["id"] not in existing_ids]
        patch[key] = new_only
    MERGE_OUT.write_text(json.dumps(patch, ensure_ascii=False, indent=2), encoding="utf-8")
    summary = " | ".join(f"{k}:+{len(patch[k])}" for k in ("consultants","clients","companies","deals","activities","pupilos"))
    print(f"\n✅ merge_patch.json generado.\n   Sólo IDs nuevos: {summary}\n   Importa este fichero desde el CRM con el botón 'Importar merge'.")

# ── AUTO-GEN SCHEMA.md ────────────────────────────────
def write_schema_doc(seed):
    """Genera SCHEMA.md humano-legible con todas las entidades, campos y conteos."""
    out = HERE / "SCHEMA.md"
    lines = [
        f"# CEO Advisors CRM — Schema {SCHEMA_VERSION}",
        "",
        f"_Auto-generado por `inject_data.py` el {seed.get('_seedTs','?')}._",
        "",
        "## Entidades",
        "",
        "| Entidad | Prefijo ID | Campos | Cantidad actual |",
        "|---|---|---|---|",
    ]
    prefixes = {"consultants":"u","clients":"c","companies":"co","deals":"d","activities":"a","pupilos":"p"}
    for k, fields in SCHEMA_FIELDS.items():
        n = len(seed.get(k, []))
        lines.append(f"| `{k}` | `{prefixes.get(k,'?')}` | {', '.join('`'+f+'`' for f in fields)} | {n} |")
    lines += [
        "",
        "## Notas de campos especiales",
        "",
        "- `consultants` tienen `passwordHash`, `passwordSalt`, `passwordIters`, `passwordMustChange` adicionales (PBKDF2-HMAC-SHA256).",
        "- `pupilos` tienen `docs[]` con `{name, path, mime, uploadedAt}` o `{name, dataUrl, ...}`.",
        "- `clients` y `companies` y `deals` también soportan `docs[]` y `comments[]`.",
        "- `deals` tienen `splits[]` con `{u: consultantId, pct: number}` que deben sumar 100.",
        "- `deals` tienen `stageHistory[]` con `{stage, ts}` para tracking.",
        "- `activities` tienen `done` y `assignedTo` (consultantId).",
        "",
        "## Reglas de validación (`inject_data.py`)",
        "",
        "- IDs únicos por entidad (sin duplicados).",
        "- Toda referencia (`clientId`, `companyId`, `mentor`, `consultantId`, `createdBy`, `splits[].u`) debe existir.",
        "- `splits` deben sumar exactamente 100.",
        "- Passwords débiles (lista WEAK_PASSWORDS) marcan `passwordMustChange:true`.",
        "",
        "## Migraciones (en HTML, `MIGRATIONS` array)",
        "",
        "Cada migración tiene `{to: N, fn}`. Se aplican en cadena si el `_schemaApplied` del DB guardado es menor que `to`. **Bumps de schema requieren** añadir nueva entrada al array, no modificar las anteriores.",
    ]
    out.write_text("\n".join(lines), encoding="utf-8")

# ── PATCH HTML (modo seed) ────────────────────────────
def patch_html(seed):
    html = CRM_HTML.read_text(encoding="utf-8")
    seed_json = json.dumps(seed, ensure_ascii=False)
    new_body = (
        "function seedData(){\n"
        f"  const d={seed_json};\n"
        "  return d;\n"
        "}"
    )
    patched, count = re.subn(r"function seedData\(\)\{[\s\S]+?\n\}(?=\n)", new_body, html)
    if count == 0:
        print("⚠ No encontré seedData() en el HTML."); sys.exit(1)
    OUTPUT_HTML.write_text(patched, encoding="utf-8")

# ── MAIN ──────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--merge", action="store_true", help="Generar merge_patch.json en lugar de sobrescribir el seed.")
    ap.add_argument("--from", dest="from_path", help="Ruta al export JSON del CRM (requerido con --merge).")
    args = ap.parse_args()

    print(f"CEO Advisors CRM — Inject Data (schema {SCHEMA_VERSION})\n")
    consultants, clients, companies, deals, activities, pupilos = parse_workbook()

    print(f"  Parseado: {len(consultants)} consultants, {len(clients)} clients, "
          f"{len(companies)} companies, {len(deals)} deals, {len(activities)} activities, {len(pupilos)} pupilos\n")

    errors, warnings = validate(consultants, clients, companies, deals, activities, pupilos)
    if warnings:
        print("WARNINGS:")
        for w in warnings: print("  ⚠", w)
        print()
    if errors:
        print("ERRORES — corrige antes de inyectar:")
        for e in errors: print("  ✗", e)
        print(f"\n✗ {len(errors)} error(es). Abortando.")
        sys.exit(2)
    print("  ✓ Validación pre-flight OK\n")

    seed = build_seed(consultants, clients, companies, deals, activities, pupilos)

    if args.merge:
        if not args.from_path:
            print("⚠ --merge requiere --from <export.json>."); sys.exit(1)
        merge_mode(seed, args.from_path)
        return

    patch_html(seed)
    write_schema_doc(seed)
    print(f"""✅ Production CRM generado.
   Output : {OUTPUT_HTML.name}
   ───────────────────────────────────
   Consultants : {len(consultants)}
   Clients     : {len(clients)}
   Companies   : {len(companies)}
   Deals       : {len(deals)}
   Activities  : {len(activities)}
   Pupilos     : {len(pupilos)}
   ──────────────────────────────────
   First login : {consultants[0]['name']} / {consultants[0]['email']}
""")

if __name__ == "__main__":
    main()
