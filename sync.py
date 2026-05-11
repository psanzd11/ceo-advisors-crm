#!/usr/bin/env python3
"""
CEO Advisors CRM — Sync round-trip
Schema v2026.05.1

Toma un export JSON del CRM (botón "Exportar JSON") y:
  1. Actualiza CEO_Advisors_CRM_DataTemplate_v2.xlsx con los datos más recientes
     (match por ID; si no, fuzzy-match por nombre; si no, append como fila nueva).
  2. Regenera CEO_Advisors_CRM_PRODUCTION.html ejecutando inject_data.py.

Uso:
  py sync.py ceoadvisors_crm_export.json
  py sync.py path\\to\\export.json --no-regen   # sólo actualiza el xlsx
  py sync.py --skip-xlsx export.json            # sólo regenera el HTML desde el JSON

Si el xlsx está abierto en Excel, se guarda como _v2.NEW.xlsx y el script avisa.
"""
import json, sys, re, shutil, subprocess, argparse, unicodedata
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl…")
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl

HERE = Path(__file__).parent
TEMPLATE = HERE / "CEO_Advisors_CRM_DataTemplate_v2.xlsx"
INJECT_SCRIPT = HERE / "inject_data.py"
PUPILO_DOCS_DIR = "pupilo_docs"
PROCESSED_DIR = HERE / "processed_exports"
PROCESSED_KEEP = 10  # snapshots a conservar; el resto se borra automáticamente

# ── Fuzzy name match (handles typos like Salame/Salome, Travesario/Traversari) ──
def norm(s):
    if not s: return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9 -]", "", s).strip()

def tokens(s):
    return [t for t in re.split(r"[ \-]+", norm(s)) if t]

def lev(a, b):
    if a == b: return 0
    if not a: return len(b)
    if not b: return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a):
        curr = [i + 1]
        for j, cb in enumerate(b):
            cost = 0 if ca == cb else 1
            curr.append(min(curr[j] + 1, prev[j + 1] + 1, prev[j] + cost))
        prev = curr
    return prev[-1]

def close(x, y, max_dist=2):
    if x == y: return True
    if not x or not y: return False
    if x.startswith(y) or y.startswith(x): return True
    if x in y or y in x: return True
    if abs(len(x) - len(y)) > max_dist + 1: return False
    return lev(x, y) <= max_dist

def fuzzy_eq(a, b):
    ta, tb = tokens(a), tokens(b)
    if not ta or not tb: return False
    if not close(ta[0], tb[0], 2): return False
    rest_a = ta[1:] or ta
    rest_b = tb[1:] or tb
    return any(close(x, y, 2) for x in rest_a for y in rest_b)

# ── Sheet schema (column index, value-getter from entity) ──
SHEETS = {
    "👥 Consultants": {
        "key": "consultants", "prefix": "u",
        "cols": [
            ("Full Name", lambda e: e.get("name", "")),
            ("Role", lambda e: e.get("role", "")),
            ("Is CEO?", lambda e: "Yes" if e.get("isCEO") else "No"),
            ("Is Admin?", lambda e: "Yes" if e.get("isAdmin") else "No"),
            ("Email", lambda e: e.get("email", "")),
            ("Password", lambda e: e.get("password", "")),
            ("Region", lambda e: e.get("region", "")),
            ("Bio", lambda e: e.get("bio", "")),
        ],
    },
    "🤝 Clients": {
        "key": "clients", "prefix": "c",
        "cols": [
            ("Full Name", lambda e: e.get("name", "")),
            ("Email", lambda e: e.get("email", "")),
            ("Phone", lambda e: e.get("phone", "")),
            ("Country", lambda e: e.get("country", "")),
            ("City", lambda e: e.get("city", "")),
            ("Net Worth (USD)", lambda e: e.get("netWorth") or ""),
            ("Lead Source", lambda e: e.get("source", "")),
            ("Tier", lambda e: e.get("tier", "")),
            ("Notes", lambda e: e.get("notes", "")),
        ],
    },
    "🏢 Companies": {
        "key": "companies", "prefix": "co",
        "cols": [
            ("Company Name", lambda e: e.get("name", "")),
            ("Industry", lambda e: e.get("industry", "")),
            ("Country", lambda e: e.get("country", "")),
            ("Employees", lambda e: e.get("employees") or ""),
            ("Net Worth (USD)", lambda e: e.get("netWorth") or ""),
            ("Website", lambda e: e.get("website", "")),
            ("Linked Client ID(s)", lambda e: ", ".join(e.get("clientIds", [])) if isinstance(e.get("clientIds"), list) else (e.get("clientIds") or "")),
            ("Notes", lambda e: e.get("notes", "")),
        ],
    },
    "💼 Deals": {
        "key": "deals", "prefix": "d",
        "cols": [
            ("Deal Title", lambda e: e.get("title", "")),
            ("Client ID", lambda e: e.get("clientId", "")),
            ("Company ID", lambda e: e.get("companyId", "")),
            ("Type", lambda e: e.get("type", "")),
            ("Stage", lambda e: e.get("stage", "")),
            ("Amount (USD)", lambda e: e.get("amount") or ""),
            ("Expected Close Date", lambda e: e.get("closeDate", "")),
            ("Created Date", lambda e: e.get("createdAt", "")),
            ("Advisor 1 ID", lambda e: (e.get("splits") or [{}])[0].get("u", "")),
            ("Advisor 1 %", lambda e: (e.get("splits") or [{}])[0].get("pct", "")),
            ("Advisor 2 ID", lambda e: (e.get("splits") or [{}, {}])[1].get("u", "") if len(e.get("splits") or []) >= 2 else ""),
            ("Advisor 2 %", lambda e: (e.get("splits") or [{}, {}])[1].get("pct", "") if len(e.get("splits") or []) >= 2 else ""),
            ("Advisor 3 ID", lambda e: (e.get("splits") or [{}, {}, {}])[2].get("u", "") if len(e.get("splits") or []) >= 3 else ""),
            ("Advisor 3 %", lambda e: (e.get("splits") or [{}, {}, {}])[2].get("pct", "") if len(e.get("splits") or []) >= 3 else ""),
            ("Notes", lambda e: e.get("notes", "")),
        ],
    },
    "📋 Activities": {
        "key": "activities", "prefix": "a",
        "cols": [
            ("Type", lambda e: e.get("type", "")),
            ("Date", lambda e: e.get("date", "")),
            ("Client ID", lambda e: e.get("clientId", "")),
            ("Company ID", lambda e: e.get("companyId", "")),
            ("Deal ID", lambda e: e.get("dealId", "")),
            ("Title", lambda e: e.get("title", "")),
            ("Notes", lambda e: e.get("notes", "")),
            ("Done?", lambda e: "Yes" if e.get("done") else "No"),
            ("Created By", lambda e: e.get("createdBy", "")),
        ],
    },
    "🎓 Pupilos": {
        "key": "pupilos", "prefix": "p",
        "cols": [
            ("Full Name", lambda e: e.get("name", "")),
            ("Email", lambda e: e.get("email", "")),
            ("University", lambda e: e.get("university", "")),
            ("Program", lambda e: e.get("program", "")),
            ("Start Date", lambda e: e.get("startDate", "")),
            ("End Date", lambda e: e.get("endDate", "")),
            ("Mentor ID", lambda e: e.get("mentor", "")),
            ("Region", lambda e: e.get("region", "")),
            ("Consultant ID", lambda e: e.get("consultantId", "")),
            ("Left Company", lambda e: e.get("leftCompany", "")),
            ("Left Role", lambda e: e.get("leftRole", "")),
            ("Notes", lambda e: e.get("notes", "")),
            ("CV File", lambda e: ((e.get("docs") or [{}])[0].get("name", "")) if e.get("docs") else ""),
        ],
    },
}

def name_field(entity, key):
    """Para fuzzy-match: nombre principal según el tipo."""
    return entity.get("name") or entity.get("title") or ""

def archive_json(json_path):
    """Mueve el JSON procesado a processed_exports/ con timestamp.
    Mantiene sólo los últimos PROCESSED_KEEP snapshots."""
    json_path = Path(json_path)
    if not json_path.exists():
        return None
    PROCESSED_DIR.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    dst = PROCESSED_DIR / f"{stamp}_{json_path.name}"
    try:
        shutil.move(str(json_path), dst)
    except Exception as e:
        print(f"  ⚠ No pude archivar {json_path.name}: {e}")
        return None
    # Rotación: borrar los más antiguos
    files = sorted(PROCESSED_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    for old in files[PROCESSED_KEEP:]:
        try: old.unlink()
        except: pass
    return dst

# ── SYNC ──
def sync_xlsx(data):
    if not TEMPLATE.exists():
        print(f"✗ No encuentro {TEMPLATE.name}. Aborto.")
        sys.exit(1)

    # Backup
    backup = TEMPLATE.with_suffix(f".bak-{datetime.now():%Y%m%d-%H%M%S}.xlsx")
    shutil.copy2(TEMPLATE, backup)

    wb = openpyxl.load_workbook(TEMPLATE)
    summary = {}
    conflicts = []  # 2.3: campos donde JSON y Excel discrepan; el usuario decide después

    for sheet_name, cfg in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠ {sheet_name}: hoja no encontrada — skip")
            continue
        ws = wb[sheet_name]
        key = cfg["key"]
        prefix = cfg["prefix"]
        cols = cfg["cols"]
        entities = data.get(key, []) or []

        # Lectura del estado actual (filas 4+)
        existing_by_id = {}
        existing_rows = []
        for r in range(4, ws.max_row + 1):
            current_id = ws.cell(r, 1).value
            current_name = ws.cell(r, 2).value
            if not current_name:
                continue
            # Si la celda A es fórmula sin calcular, derivar el ID por convención
            if isinstance(current_id, str) and current_id.startswith("="):
                current_id = f"{prefix}{r-3}"
            existing_rows.append({"row": r, "id": current_id or f"{prefix}{r-3}", "name": current_name})
            if current_id:
                existing_by_id[current_id] = r

        used_rows = set()
        updated, appended = 0, 0

        for e in entities:
            eid = e.get("id", "")
            target_row = None

            # 1. Match por ID
            if eid in existing_by_id and existing_by_id[eid] not in used_rows:
                target_row = existing_by_id[eid]

            # 2. Fuzzy match por nombre (sólo entidades con nombre)
            if target_row is None:
                ename = name_field(e, key)
                if ename:
                    for er in existing_rows:
                        if er["row"] in used_rows:
                            continue
                        if fuzzy_eq(ename, er["name"]):
                            target_row = er["row"]
                            break

            # 3. Append como nueva fila
            if target_row is None:
                target_row = max((er["row"] for er in existing_rows), default=3) + 1
                # Mantén la fórmula de auto-id en col A
                ws.cell(target_row, 1).value = f'=IF(B{target_row}="","","{prefix}"&ROW()-3)'
                existing_rows.append({"row": target_row, "id": f"{prefix}{target_row-3}", "name": e.get("name") or e.get("title") or ""})
                appended += 1
            else:
                updated += 1

            used_rows.add(target_row)

            # NON-DESTRUCTIVE MERGE + CONFLICT DETECTION:
            #   - Si JSON vacío y Excel lleno → preserva Excel.
            #   - Si JSON lleno y Excel lleno con valor distinto → CONFLICTO (anota, NO pisa).
            #   - Si JSON lleno y Excel vacío/igual → escribe.
            def _is_empty(v):
                return v is None or (isinstance(v, str) and not v.strip())
            def _norm(v):
                if v is None: return ""
                return str(v).strip()
            for i, (header, getter) in enumerate(cols):
                col = 2 + i
                new_val = getter(e)
                cur_val = ws.cell(target_row, col).value
                if _is_empty(new_val) and not _is_empty(cur_val):
                    continue  # preserva
                if not _is_empty(new_val) and not _is_empty(cur_val) and _norm(new_val) != _norm(cur_val):
                    # CONFLICTO — registrar y NO pisar (el usuario decide manualmente)
                    conflicts.append({
                        "sheet": sheet_name, "row": target_row,
                        "id": e.get("id", ""), "name": e.get("name") or e.get("title") or "",
                        "field": header,
                        "excel": _norm(cur_val), "json": _norm(new_val),
                    })
                    continue
                ws.cell(target_row, col).value = new_val

        summary[sheet_name] = (updated, appended)

    # Guardar
    try:
        wb.save(TEMPLATE)
        print(f"\n✓ Plantilla actualizada: {TEMPLATE.name}")
        print(f"  Backup: {backup.name}")
    except PermissionError:
        alt = TEMPLATE.with_suffix(".NEW.xlsx")
        wb.save(alt)
        print(f"\n⚠ {TEMPLATE.name} está abierto en Excel.")
        print(f"  → Cambios guardados en {alt.name} — ciérralo y renómbralo manualmente.")
        return False

    print("\n  Resumen por hoja:")
    for s, (u, a) in summary.items():
        print(f"    {s:25s}  ↻ {u} actualizados · ＋ {a} nuevos")

    if conflicts:
        cf_path = HERE / "conflicts.json"
        cf_path.write_text(json.dumps(conflicts, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"\n  ⚠ {len(conflicts)} CONFLICTOS detectados (Excel y JSON discrepan).")
        print(f"     Detalle en {cf_path.name}. El Excel NO se sobrescribe en esos campos.")
        print(f"     Para resolver: abre conflicts.json, decide manualmente cuál vale, edita el Excel y vuelve a ejecutar.")
        # Mostrar primeros 5 conflictos en consola
        for c in conflicts[:5]:
            print(f"       · {c['sheet']} fila {c['row']} ({c['name']}) campo '{c['field']}': excel='{c['excel'][:40]}' vs json='{c['json'][:40]}'")
        if len(conflicts) > 5:
            print(f"       ... y {len(conflicts)-5} más en {cf_path.name}")

    return True

# ── REGEN HTML ──
def regen_html():
    if not INJECT_SCRIPT.exists():
        print(f"✗ No encuentro {INJECT_SCRIPT.name}. Saltando regeneración.")
        return
    print(f"\n→ Ejecutando {INJECT_SCRIPT.name}…")
    r = subprocess.run([sys.executable, str(INJECT_SCRIPT)], cwd=HERE)
    if r.returncode != 0:
        print(f"⚠ inject_data.py retornó {r.returncode}.")

# ── MAIN ──
def main():
    ap = argparse.ArgumentParser(description="Sync CRM JSON export to xlsx template + production HTML")
    ap.add_argument("json_path", help="Ruta al ceoadvisors_crm_export.json descargado del CRM")
    ap.add_argument("--no-regen", action="store_true", help="No regenerar el HTML de producción")
    ap.add_argument("--skip-xlsx", action="store_true", help="Saltar actualización del xlsx, sólo regenerar HTML")
    args = ap.parse_args()

    p = Path(args.json_path)
    if not p.exists():
        print(f"X No encuentro {p}. Aborto.")
        sys.exit(1)

    data = json.loads(p.read_text(encoding="utf-8"))
    schema = data.get("schemaVersion", "?")
    print(f"CEO Advisors CRM - sync round-trip")
    print(f"  Origen   : {p.name}  (schema {schema})")
    print(f"  Plantilla: {TEMPLATE.name}\n")

    counts = {k: len(data.get(k, [])) for k in ["consultants", "clients", "companies", "deals", "activities", "pupilos"]}
    print("  En el JSON: " + " | ".join(f"{k}:{v}" for k, v in counts.items()))

    if not args.skip_xlsx:
        ok = sync_xlsx(data)
        if not ok:
            print("\n-> No regenero el HTML porque el xlsx no se pudo guardar.")
            return

    if not args.no_regen:
        regen_html()

    archived = archive_json(p)
    if archived:
        print(f"\n  JSON archivado en {archived.relative_to(HERE)}")
        print(f"     (se conservan los ultimos {PROCESSED_KEEP} snapshots como backup)")

if __name__ == "__main__":
    main()
